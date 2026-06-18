"""Publish the yield workbook to Google Sheets (gspread).

Google Sheets is the live dashboard: the team views it over the company network
and types into the Issue Log ``Action`` / ``Status`` columns. This module pushes
the freshly-computed sheets up there **without clobbering those edits** — it
reads the existing Action/Status back first (keyed by the hidden ``_key``
column) so the caller can fold them into the rebuild.

Not exercised in the sandbox (no network to Google). Needs ``gspread`` and a
Google service-account JSON key shared with the target spreadsheet. See
README for setup.

Typical flow (see ops_yield_cli.py --gsheet-url)::

    sh    = connect(creds_path, sheet_url)
    prior = read_actions(sh)                       # pull back user edits
    write_workbook(records, local_xlsx, prior_actions=prior)
    push_workbook(sh, local_xlsx)                  # publish every tab
"""

from __future__ import annotations

from openpyxl import load_workbook


def connect(creds_path: str, sheet_url: str):
    """Open the target spreadsheet via a service-account key."""
    import gspread  # imported lazily so the package works without it

    gc = gspread.service_account(filename=creds_path)
    return gc.open_by_url(sheet_url)


def read_actions(spreadsheet) -> dict:
    """Return {issue_key: (action, status)} from the live Issue Log tab."""
    try:
        ws = spreadsheet.worksheet("Issue Log")
    except Exception:
        return {}
    values = ws.get_all_values()
    if not values:
        return {}
    headers = values[0]
    try:
        ik, ia, ist = (headers.index("_key"),
                       headers.index("Action"), headers.index("Status"))
    except ValueError:
        return {}
    out = {}
    for row in values[1:]:
        if len(row) <= max(ik, ia, ist):
            continue
        key, action, status = row[ik], row[ia], row[ist]
        if key and (action or status):
            out[key] = (action, status)
    return out


def read_messages(spreadsheet, sheet_name: str = "Messages",
                  text_col: str = "text") -> str:
    """Pull the LINE messages collector.gs logged into the ``Messages`` tab and
    return them as one blob (blank-line separated) for ``parse_messages``."""
    ws = spreadsheet.worksheet(sheet_name)
    rows = ws.get_all_records()  # first row = headers: timestamp, source, text
    texts = [str(r.get(text_col, "")).strip() for r in rows]
    return "\n\n".join(t for t in texts if t)


def _sheet_values(ws) -> list[list]:
    """openpyxl worksheet -> 2D list of JSON-safe values for gspread."""
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append(["" if v is None else v for v in row])
    return out


# Tabs that get a native "All yield trend" line chart embedded next to the table.
_CHART_TABS = ("Weekly Yield", "Monthly Yield", "Quarterly Yield")


def push_workbook(spreadsheet, xlsx_path: str) -> list[str]:
    """Overwrite each tab of ``spreadsheet`` with the xlsx's sheets."""
    wb = load_workbook(xlsx_path)
    pushed = []
    line_specs: dict[str, tuple[int, int]] = {}
    dash_spec = None
    for name in wb.sheetnames:
        values = _sheet_values(wb[name])
        rows = max(len(values), 1)
        cols = max((len(r) for r in values), default=1)
        try:
            ws = spreadsheet.worksheet(name)
            ws.clear()
        except Exception:
            ws = spreadsheet.add_worksheet(title=name, rows=rows + 10, cols=cols + 2)
        if values:
            ws.update(values, value_input_option="USER_ENTERED")
        if name in _CHART_TABS and len(values) > 1:
            line_specs[name] = (len(values), cols)
        if name == "Dashboard":
            sx = wb["Dashboard"]
            m_rows = sum(1 for r in range(2, sx.max_row + 1) if sx.cell(r, 1).value not in (None, ""))
            d_rows = sum(1 for r in range(2, sx.max_row + 1) if sx.cell(r, 6).value not in (None, ""))
            dash_spec = (m_rows, d_rows)
        pushed.append(name)
    _rebuild_charts(spreadsheet, line_specs, dash_spec)
    return pushed


def _src(sid, r0, r1, c0, c1):
    return {"sheetId": sid, "startRowIndex": r0, "endRowIndex": r1,
            "startColumnIndex": c0, "endColumnIndex": c1}


def _series(sid, r1, col, ctype, axis):
    """One chart series = a single column ``col`` (0-based), rows 0..r1."""
    return {"series": {"sourceRange": {"sources": [_src(sid, 0, r1, col, col + 1)]}},
            "type": ctype, "targetAxis": axis}


def _rebuild_charts(spreadsheet, line_specs: dict, dash_spec) -> None:
    """(Re)create every embedded chart. ``ws.clear()`` wipes values but leaves
    charts behind, so we delete existing charts on the touched tabs first.

    * ``line_specs`` maps a period tab -> (n_rows, n_cols): a LINE chart of the
      last column (รวม/All) over column A.
    * ``dash_spec`` = (m_rows, d_rows) for the Dashboard tab: a COMBO of
      clustered Pass/Fail columns + a Yield line, and a COMBO of stacked
      root-cause columns + %defect-rate lines.
    """
    meta = spreadsheet.fetch_sheet_metadata(
        params={"fields": "sheets(properties(sheetId,title),charts(chartId))"})
    by_title = {}
    for sh in meta.get("sheets", []):
        props = sh.get("properties", {})
        by_title[props.get("title")] = (
            props.get("sheetId"),
            [c.get("chartId") for c in sh.get("charts", [])],
        )

    requests: list[dict] = []

    def _wipe(name):
        if name in by_title:
            for cid in by_title[name][1]:
                requests.append({"deleteEmbeddedObject": {"objectId": cid}})

    # --- period line charts -------------------------------------------------
    for name, (n_rows, n_cols) in line_specs.items():
        if name not in by_title:
            continue
        sid = by_title[name][0]
        _wipe(name)
        requests.append({"addChart": {"chart": {
            "spec": {
                "title": f"{name} — รวม (All) yield trend",
                "basicChart": {
                    "chartType": "LINE", "legendPosition": "NO_LEGEND",
                    "axis": [{"position": "LEFT_AXIS", "title": "Yield"}],
                    "domains": [{"domain": {"sourceRange": {
                        "sources": [_src(sid, 0, n_rows, 0, 1)]}}}],
                    "series": [{"targetAxis": "LEFT_AXIS", "series": {"sourceRange": {
                        "sources": [_src(sid, 0, n_rows, n_cols - 1, n_cols)]}}}],
                    "headerCount": 1,
                }},
            "position": {"overlayPosition": {"anchorCell": {
                "sheetId": sid, "rowIndex": 1, "columnIndex": n_cols + 1}}},
        }}})

    # --- Dashboard combo charts --------------------------------------------
    if dash_spec and "Dashboard" in by_title:
        sid = by_title["Dashboard"][0]
        m_rows, d_rows = dash_spec
        _wipe("Dashboard")
        t1, t2 = m_rows + 1, d_rows + 1  # endRowIndex (header + data)
        # 1) per-machine pass/fail columns + yield line
        requests.append({"addChart": {"chart": {
            "spec": {
                "title": "Per-machine pass/fail + yield (latest day)",
                "basicChart": {
                    "chartType": "COMBO", "legendPosition": "BOTTOM_LEGEND",
                    "stackedType": "NOT_STACKED",
                    "axis": [{"position": "BOTTOM_AXIS"},
                             {"position": "LEFT_AXIS", "title": "Units"},
                             {"position": "RIGHT_AXIS", "title": "Yield"}],
                    "domains": [{"domain": {"sourceRange": {
                        "sources": [_src(sid, 0, t1, 0, 1)]}}}],
                    "series": [_series(sid, t1, 1, "COLUMN", "LEFT_AXIS"),
                               _series(sid, t1, 2, "COLUMN", "LEFT_AXIS"),
                               _series(sid, t1, 3, "LINE", "RIGHT_AXIS")],
                    "headerCount": 1,
                }},
            "position": {"overlayPosition": {"anchorCell": {
                "sheetId": sid, "rowIndex": m_rows + 2, "columnIndex": 0}}},
        }}})
        # 2) stacked root-cause columns + %defect-rate lines
        requests.append({"addChart": {"chart": {
            "spec": {
                "title": "Root-cause tracker — fail units & %defect rate (MA7)",
                "basicChart": {
                    "chartType": "COMBO", "legendPosition": "BOTTOM_LEGEND",
                    "stackedType": "STACKED",
                    "axis": [{"position": "BOTTOM_AXIS"},
                             {"position": "LEFT_AXIS", "title": "Fail units"},
                             {"position": "RIGHT_AXIS", "title": "% defect (MA7)"}],
                    "domains": [{"domain": {"sourceRange": {
                        "sources": [_src(sid, 0, t2, 5, 6)]}}}],
                    "series": [_series(sid, t2, 6, "COLUMN", "LEFT_AXIS"),
                               _series(sid, t2, 7, "COLUMN", "LEFT_AXIS"),
                               _series(sid, t2, 8, "COLUMN", "LEFT_AXIS"),
                               _series(sid, t2, 9, "LINE", "RIGHT_AXIS"),
                               _series(sid, t2, 10, "LINE", "RIGHT_AXIS"),
                               _series(sid, t2, 11, "LINE", "RIGHT_AXIS")],
                    "headerCount": 1,
                }},
            "position": {"overlayPosition": {"anchorCell": {
                "sheetId": sid, "rowIndex": m_rows + 20, "columnIndex": 0}}},
        }}})

    if requests:
        spreadsheet.batch_update({"requests": requests})
