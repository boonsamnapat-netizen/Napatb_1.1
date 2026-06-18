"""Write parsed production records into an Excel workbook.

Sheets:
  * ``Records``         — one row per LINE message (the audit trail).
  * ``Daily Yield``     — matrix: rows = date, columns = each machine + รวม (All).
  * ``Weekly Yield``    — matrix: rows = Sat→Fri work week (see ``fiscal``).
  * ``Monthly Yield``   — matrix: rows = calendar month.
  * ``Quarterly Yield`` — matrix: rows = fiscal quarter (Q1 FY27 = 4 Jul 2026).
  * ``Issue Log``       — time-series of reports that had a fail, with editable
                          ``Action`` / ``Status`` columns the user fills in.
  * ``By Shift`` / ``By Operator`` / ``Fail Causes`` — summaries.

Yield cells aggregate by **summing pass and total** across every report for
that machine/period, then dividing — this is the correct way to combine
yields, not by averaging percentages. A green→red colour scale makes low
yield pop.

Re-running appends rather than clobbers: existing ``Records`` are read back
(idempotent on ``raw`` text), and any ``Action`` / ``Status`` the user typed in
``Issue Log`` is preserved by matching a stable per-report key.
"""

from __future__ import annotations

import hashlib
import os
from collections import defaultdict
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import fiscal
from .analysis import by_operator, by_shift, cause_summary, classify_causes
from .parser import Record

RECORD_HEADERS = [
    "Date", "Week", "Machine", "Mode", "Issue Type", "Shift",
    "Total", "Fail L1", "Fail L2", "Splash", "Pass", "Yield",
    "Status Note", "Warnings", "Raw", "Sender",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_PCT = "0.0%"


def _machine_key(m):
    """Natural sort: numeric machines first (1,2,..10), text after."""
    s = str(m)
    return (0, int(s)) if s.isdigit() else (1, s)


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _read_existing_records(path: str) -> list[dict]:
    if not os.path.exists(path):
        return []
    wb = load_workbook(path)
    if "Records" not in wb.sheetnames:
        return []
    ws = wb["Records"]
    rows = []
    headers = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(headers, r)))
    return rows


def _record_to_row(rec: Record) -> list:
    d = rec.work_date
    return [
        d.isoformat() if d else "",
        fiscal.week(d)[1] if d else "",
        rec.machine,
        rec.mode,
        rec.issue_type,
        rec.shift,
        rec.total,
        rec.fail_l1,
        rec.fail_l2,
        rec.splash,
        rec.passed,
        rec.yield_ratio,
        rec.status_note,
        "; ".join(rec.warnings),
        rec.raw,
        rec.sender,
    ]


def _row_key(row: list) -> tuple:
    """Dedup key: same date+machine+raw text => same report."""
    return (row[0], row[2], row[14])


def _issue_key(row: list) -> str:
    """Stable string key for an Issue-Log row (survives regeneration)."""
    raw = str(row[14] or "")
    h = hashlib.md5(raw.encode("utf-8")).hexdigest()[:8]
    return f"{row[0]}|{row[2]}|{row[5]}|{h}"


def _read_existing_actions(path: str) -> dict:
    """Map issue-key -> (Action, Status) from a prior Issue Log, to preserve
    what the user typed when we rebuild the workbook."""
    if not os.path.exists(path):
        return {}
    wb = load_workbook(path)
    if "Issue Log" not in wb.sheetnames:
        return {}
    ws = wb["Issue Log"]
    headers = [c.value for c in ws[1]]
    try:
        ik = headers.index("_key")
        ia = headers.index("Action")
        ist = headers.index("Status")
    except ValueError:
        return {}
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        key = r[ik]
        if key and (r[ia] or r[ist]):
            out[key] = (r[ia] or "", r[ist] or "")
    return out


ISSUE_HEADERS = [
    "Date", "Week", "Machine", "Shift", "Total", "Fail L1", "Fail L2",
    "Splash", "Fails", "Yield", "Cause", "Sender", "Note",
    "Action", "Status", "_key",
]
_STATUS_CHOICES = '"Open,In progress,Done,Ignore"'


def write_workbook(records: list[Record], path: str, *, append: bool = True,
                   prior_actions: dict | None = None) -> dict:
    """Build/refresh the workbook at ``path`` from ``records``.

    ``prior_actions`` (key -> (action, status)) lets a caller inject the
    Action/Status a user typed elsewhere (e.g. in Google Sheets) so it survives
    the rebuild; it is merged on top of any actions already in ``path``.
    Returns a small summary dict (counts) for logging.
    """
    new_rows = [_record_to_row(r) for r in records]

    existing = []
    if append:
        for er in _read_existing_records(path):
            existing.append([er.get(h) for h in RECORD_HEADERS])

    seen = set()
    all_rows = []
    for row in existing + new_rows:
        k = _row_key(row)
        if k in seen:
            continue
        seen.add(k)
        all_rows.append(row)

    # sort by date then machine for a tidy sheet
    all_rows.sort(key=lambda r: (str(r[0]), _machine_key(r[2])))

    # facts: (shift, total, passed, sender, notes)
    facts = [(r[5], r[6] or 0, r[10] or 0, r[15], r[12]) for r in all_rows]

    actions = _read_existing_actions(path) if append else {}
    if prior_actions:
        actions.update(prior_actions)

    wb = Workbook()
    _build_records_sheet(wb, all_rows)
    machines = _build_matrix_sheet(wb, all_rows, "Daily Yield",
                                   period_func=fiscal.day, key_title="Date")
    _build_matrix_sheet(wb, all_rows, "Weekly Yield",
                        period_func=fiscal.week, key_title="Work week (Sat–Fri)",
                        chart=True)
    _build_matrix_sheet(wb, all_rows, "Monthly Yield",
                        period_func=fiscal.month, key_title="Month", chart=True)
    _build_matrix_sheet(wb, all_rows, "Quarterly Yield",
                        period_func=fiscal.quarter, key_title="Quarter (FY)",
                        chart=True)
    _build_dashboard_sheet(wb, all_rows)
    n_issues = _build_issue_log(wb, all_rows, actions)
    _build_shift_sheet(wb, facts)
    _build_operator_sheet(wb, facts)
    _build_cause_sheet(wb, facts)

    wb.save(path)
    return {
        "records_total": len(all_rows),
        "records_added": len(new_rows),
        "machines": sorted(machines),
        "issues": n_issues,
        "path": path,
    }


def _build_records_sheet(wb: Workbook, rows: list[list]) -> None:
    ws = wb.active
    ws.title = "Records"
    ws.append(RECORD_HEADERS)
    yield_col = RECORD_HEADERS.index("Yield") + 1
    for row in rows:
        ws.append(row)
        ws.cell(row=ws.max_row, column=yield_col).number_format = _PCT
    _style_header(ws, len(RECORD_HEADERS))
    widths = [12, 9, 8, 8, 12, 7, 7, 8, 8, 8, 7, 8, 24, 22, 40, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if ws.max_row > 1:
        ws.conditional_formatting.add(
            f"L2:L{ws.max_row}",
            ColorScaleRule(
                start_type="num", start_value=0.8, start_color="F8696B",
                mid_type="num", mid_value=0.95, mid_color="FFEB84",
                end_type="num", end_value=1.0, end_color="63BE7B",
            ),
        )


def _build_matrix_sheet(
    wb: Workbook, rows: list[list], title: str, *, period_func, key_title: str,
    chart: bool = False,
) -> set:
    """Aggregate pass/total into a period x machine yield matrix.

    ``period_func(date) -> (sort_key, label)`` groups each record's date into a
    period (day/week/month/quarter). Rows sort by ``sort_key`` so chronological
    order holds even when labels (e.g. "Q1 FY28") wouldn't sort alphabetically.
    """
    # agg[sort_key][machine] = [pass_sum, total_sum]; labels[sort_key] = label
    agg: dict = defaultdict(lambda: defaultdict(lambda: [0, 0]))
    labels: dict = {}
    machines: set = set()
    for row in rows:
        ds = row[0]
        machine = row[2]
        if not ds or not machine:
            continue
        d = date.fromisoformat(str(ds)[:10])
        total = row[6] or 0
        passed = row[10] or 0
        sort_key, label = period_func(d)
        labels[sort_key] = label
        machines.add(machine)
        agg[sort_key][machine][0] += passed
        agg[sort_key][machine][1] += total
        agg[sort_key]["ALL"][0] += passed
        agg[sort_key]["ALL"][1] += total

    machine_cols = sorted(machines, key=_machine_key)
    ws = wb.create_sheet(title)
    header = [key_title] + [f"เครื่อง {m}" for m in machine_cols] + ["รวม (All)"]
    ws.append(header)

    for sk in sorted(agg):
        line = [labels[sk]]
        for m in machine_cols:
            p, t = agg[sk].get(m, [0, 0])
            line.append(p / t if t else None)
        p, t = agg[sk]["ALL"]
        line.append(p / t if t else None)
        ws.append(line)
        for c in range(2, len(header) + 1):
            ws.cell(row=ws.max_row, column=c).number_format = _PCT

    _style_header(ws, len(header))
    ws.column_dimensions["A"].width = max(12, len(key_title) + 2,
                                          *(len(str(lbl)) + 2 for lbl in labels.values())) \
        if labels else 12
    for i in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11
    if ws.max_row > 1 and len(header) > 1:
        last_col = get_column_letter(len(header))
        ws.conditional_formatting.add(
            f"B2:{last_col}{ws.max_row}",
            ColorScaleRule(
                start_type="num", start_value=0.8, start_color="F8696B",
                mid_type="num", mid_value=0.95, mid_color="FFEB84",
                end_type="num", end_value=1.0, end_color="63BE7B",
            ),
        )

    if chart and ws.max_row > 1:
        _add_trend_chart(ws, title, n_rows=ws.max_row, all_col=len(header))
    return machines


def _add_trend_chart(ws, title: str, *, n_rows: int, all_col: int) -> None:
    """Embed a line chart of the รวม (All) yield trend to the right of the table."""
    ch = LineChart()
    ch.title = f"{title} — รวม (All)"
    ch.style = 12
    ch.height = 8
    ch.width = 22
    ch.y_axis.title = "Yield"
    ch.y_axis.numFmt = "0%"
    ch.x_axis.title = ws.cell(row=1, column=1).value
    data = Reference(ws, min_col=all_col, min_row=1, max_row=n_rows)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    anchor = f"{get_column_letter(all_col + 2)}2"
    ws.add_chart(ch, anchor)


# Fixed layout for the Dashboard tab (both the openpyxl charts below and the
# Google-Sheets native charts in to_gsheet.py rely on these positions).
DASH_T1_COL = 1          # per-machine table: A..D  (Machine, Pass, Fail, Yield)
DASH_T2_COL = 6          # root-cause table:  F..L  (Date, L1, L2, Splash, 3×%MA)
DASH_ROOT_DAYS = 30      # how many recent active days the root-cause table covers
_MA_WINDOW = 7


def _moving_avg(vals: list, w: int = _MA_WINDOW) -> list:
    out = []
    for i in range(len(vals)):
        window = vals[max(0, i - w + 1):i + 1]
        out.append(sum(window) / len(window) if window else 0.0)
    return out


def _build_dashboard_sheet(wb: Workbook, rows: list[list]) -> None:
    """A visual tab: per-machine pass/fail+yield (latest day) and a root-cause
    time series, each backed by a small table and an embedded combo chart."""
    ws = wb.create_sheet("Dashboard")

    # ---- Table 1: per-machine pass/fail/yield for the latest day -----------
    dates = [str(r[0]) for r in rows if r[0]]
    latest = max(dates) if dates else ""
    agg: dict = defaultdict(lambda: [0, 0])  # machine -> [pass, fail]
    for r in rows:
        if str(r[0]) != latest or not r[2]:
            continue
        passed = r[10] or 0
        fails = (r[7] or 0) + (r[8] or 0) + (r[9] or 0)
        agg[r[2]][0] += passed
        agg[r[2]][1] += fails
    machine_cols = sorted(agg, key=_machine_key)

    ws.cell(row=1, column=DASH_T1_COL, value=f"เครื่อง ({latest})")
    ws.cell(row=1, column=DASH_T1_COL + 1, value="Pass")
    ws.cell(row=1, column=DASH_T1_COL + 2, value="Fail")
    ws.cell(row=1, column=DASH_T1_COL + 3, value="Yield")
    for i, m in enumerate(machine_cols, start=2):
        p, fl = agg[m]
        ws.cell(row=i, column=DASH_T1_COL, value=m)
        ws.cell(row=i, column=DASH_T1_COL + 1, value=p)
        ws.cell(row=i, column=DASH_T1_COL + 2, value=fl)
        c = ws.cell(row=i, column=DASH_T1_COL + 3,
                    value=(p / (p + fl) if (p + fl) else None))
        c.number_format = _PCT
    t1_last = len(machine_cols) + 1

    # ---- Table 2: root-cause counts + %defect-rate MA, last N active days --
    byday: dict = defaultdict(lambda: [0, 0, 0, 0])  # date -> [L1, L2, Splash, total]
    for r in rows:
        if not r[0]:
            continue
        d = str(r[0])
        byday[d][0] += r[7] or 0
        byday[d][1] += r[8] or 0
        byday[d][2] += r[9] or 0
        byday[d][3] += r[6] or 0
    days = sorted(byday)[-DASH_ROOT_DAYS:]
    tot = [byday[d][3] or 1 for d in days]
    ma = {j: _moving_avg([byday[d][j] / t for d, t in zip(days, tot)])
          for j in range(3)}

    heads2 = ["Date", "Fail L1", "Fail L2", "Splash",
              "L1 %def (MA7)", "L2 %def (MA7)", "Splash %def (MA7)"]
    for j, h in enumerate(heads2):
        ws.cell(row=1, column=DASH_T2_COL + j, value=h)
    for i, d in enumerate(days, start=2):
        ws.cell(row=i, column=DASH_T2_COL, value=d)
        for j in range(3):
            ws.cell(row=i, column=DASH_T2_COL + 1 + j, value=byday[d][j])
        for j in range(3):
            c = ws.cell(row=i, column=DASH_T2_COL + 4 + j, value=ma[j][i - 2])
            c.number_format = _PCT
    t2_last = len(days) + 1

    _style_header(ws, DASH_T2_COL + len(heads2) - 1)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions[get_column_letter(DASH_T2_COL)].width = 12

    # ---- Chart 1: clustered Pass/Fail columns + Yield line (2nd axis) -------
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = f"Per-machine pass/fail + yield — {latest}"
    bar.height, bar.width = 9, 20
    bar.add_data(Reference(ws, min_col=DASH_T1_COL + 1, max_col=DASH_T1_COL + 2,
                           min_row=1, max_row=t1_last), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=DASH_T1_COL, min_row=2, max_row=t1_last))
    yline = LineChart()
    yline.add_data(Reference(ws, min_col=DASH_T1_COL + 3, min_row=1, max_row=t1_last),
                   titles_from_data=True)
    yline.y_axis.axId = 200
    yline.y_axis.title = "Yield"
    yline.y_axis.numFmt = "0%"
    bar.y_axis.crosses = "max"
    bar += yline
    ws.add_chart(bar, f"A{t1_last + 3}")

    # ---- Chart 2: stacked root-cause columns + %defect MA lines (2nd axis) --
    sbar = BarChart()
    sbar.type = "col"
    sbar.grouping = "stacked"
    sbar.overlap = 100
    sbar.title = "Root-cause tracker — fail units & %defect rate (MA7)"
    sbar.height, sbar.width = 9, 24
    sbar.add_data(Reference(ws, min_col=DASH_T2_COL + 1, max_col=DASH_T2_COL + 3,
                            min_row=1, max_row=t2_last), titles_from_data=True)
    sbar.set_categories(Reference(ws, min_col=DASH_T2_COL, min_row=2, max_row=t2_last))
    rline = LineChart()
    rline.add_data(Reference(ws, min_col=DASH_T2_COL + 4, max_col=DASH_T2_COL + 6,
                             min_row=1, max_row=t2_last), titles_from_data=True)
    rline.y_axis.axId = 201
    rline.y_axis.title = "% defect (MA7)"
    rline.y_axis.numFmt = "0%"
    sbar.y_axis.crosses = "max"
    sbar += rline
    ws.add_chart(sbar, f"A{t1_last + 21}")


def _summary_rows(ws, agg: dict, key_title: str, *, sort_by_reports: bool):
    """Write a (key, reports, total, pass, fail, yield) table from an agg dict."""
    ws.append([key_title, "Reports", "Total", "Pass", "Fail", "Yield"])
    items = agg.items()
    if sort_by_reports:
        items = sorted(items, key=lambda kv: kv[1][0], reverse=True)
    else:
        items = sorted(items)
    grand = [0, 0, 0]
    yield_col = 6
    for key, (reports, total, passed) in items:
        ws.append([key, reports, total, passed, total - passed,
                   passed / total if total else None])
        ws.cell(row=ws.max_row, column=yield_col).number_format = _PCT
        grand[0] += reports
        grand[1] += total
        grand[2] += passed
    ws.append([])
    r, t, p = grand
    ws.append(["รวม (All)", r, t, p, t - p, p / t if t else None])
    ws.cell(row=ws.max_row, column=yield_col).number_format = _PCT
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    _style_header(ws, 6)
    for i, w in enumerate([18, 9, 8, 8, 8, 9], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_shift_sheet(wb: Workbook, facts) -> None:
    _summary_rows(wb.create_sheet("By Shift"), by_shift(facts), "Shift",
                  sort_by_reports=False)


def _build_operator_sheet(wb: Workbook, facts) -> None:
    _summary_rows(wb.create_sheet("By Operator"), by_operator(facts), "Operator",
                  sort_by_reports=True)


def _build_cause_sheet(wb: Workbook, facts) -> None:
    ws = wb.create_sheet("Fail Causes")
    ws.append(["สาเหตุ (Cause)", "จำนวนรายงานที่กล่าวถึง", "Fail units รวม"])
    rows = sorted(cause_summary(facts).items(), key=lambda kv: kv[1][1], reverse=True)
    for cat, (reports, fail_units) in rows:
        ws.append([cat, reports, fail_units])
    _style_header(ws, 3)
    for i, w in enumerate([24, 24, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_issue_log(wb: Workbook, rows: list[list], prior_actions: dict) -> int:
    """Chronological log of reports that had a fail, with editable Action /
    Status columns whose prior values are carried over via ``prior_actions``."""
    ws = wb.create_sheet("Issue Log")
    ws.append(ISSUE_HEADERS)

    issue_rows = [r for r in rows if (r[7] or 0) + (r[8] or 0) + (r[9] or 0) > 0]
    issue_rows.sort(key=lambda r: (str(r[0]), str(r[5]), _machine_key(r[2])))

    yield_col = ISSUE_HEADERS.index("Yield") + 1
    for r in issue_rows:
        l1, l2, sp = r[7] or 0, r[8] or 0, r[9] or 0
        fails = l1 + l2 + sp
        cause = ", ".join(classify_causes(r[12] or "")) or "—"
        key = _issue_key(r)
        action, status = prior_actions.get(key, ("", ""))
        ws.append([
            r[0], r[1], r[2], r[5], r[6], l1, l2, sp, fails, r[11],
            cause, r[15], r[12], action, status, key,
        ])
        ws.cell(row=ws.max_row, column=yield_col).number_format = _PCT

    _style_header(ws, len(ISSUE_HEADERS))
    widths = [12, 9, 8, 7, 7, 7, 7, 7, 7, 8, 20, 14, 30, 30, 12, 1]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # hide the internal matching key
    ws.column_dimensions[get_column_letter(len(ISSUE_HEADERS))].hidden = True

    if ws.max_row > 1:
        ws.conditional_formatting.add(
            f"J2:J{ws.max_row}",
            ColorScaleRule(
                start_type="num", start_value=0.8, start_color="F8696B",
                mid_type="num", mid_value=0.95, mid_color="FFEB84",
                end_type="num", end_value=1.0, end_color="63BE7B",
            ),
        )
        dv = DataValidation(type="list", formula1=_STATUS_CHOICES, allow_blank=True)
        ws.add_data_validation(dv)
        status_col = get_column_letter(ISSUE_HEADERS.index("Status") + 1)
        dv.add(f"{status_col}2:{status_col}{ws.max_row}")
    return len(issue_rows)
